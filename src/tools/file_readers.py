"""File format readers: CSV, Excel, DOCX, images."""

import contextlib
import csv
from pathlib import Path

_IMAGE_EXTENSIONS = {
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".bmp",
    ".tiff",
    ".tif",
    ".ico",
    ".svg",
}

_DOCX_EXTENSIONS = {".docx"}
_PPTX_EXTENSIONS = {".pptx"}
_CSV_EXTENSIONS = {".csv", ".tsv"}
_EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".xlsb"}
_PDF_EXTENSIONS = {".pdf"}

# Лимиты усечения табличных данных (CSV/TSV/Excel).
# При total > _TABLE_TRUNCATE_THRESHOLD печатаем header + первые
# _TABLE_HEAD_ROWS строк, затем '...', затем последние _TABLE_TAIL_ROWS.
_TABLE_HEAD_ROWS = 100
_TABLE_TAIL_ROWS = 50
_TABLE_TRUNCATE_THRESHOLD = 200
# Сколько строк реально показываем при усечении (header + head + tail).
_TABLE_SHOWN_ROWS = _TABLE_HEAD_ROWS + _TABLE_TAIL_ROWS + 1
# Жёсткий потолок чтения с диска — не материализуем больше, чем нужно,
# чтобы не словить OOM на огромных файлах. Берём чуть больше порога, чтобы
# отличить ровно-пороговый файл от усечённого.
_TABLE_READ_CAP = _TABLE_TRUNCATE_THRESHOLD + 1


def _rows_to_markdown(rows: list[list[str]], total: int) -> str:
    """Канонический рендер табличных данных в markdown.

    Принимает все строки и общее число строк. Для total > 200 печатает
    header + 100 строк данных + ... + последние 50. Иначе все.
    Используется и для CSV/TSV, и для каждого листа Excel.
    """
    if not rows:
        return "(empty)"

    truncated = total > _TABLE_TRUNCATE_THRESHOLD
    display_rows: list = (
        [*rows[: _TABLE_HEAD_ROWS + 1], None, *rows[-_TABLE_TAIL_ROWS:]]
        if truncated
        else rows
    )

    md_lines: list[str] = []
    for idx, row in enumerate(display_rows):
        if row is None:
            md_lines.append(f"| ... ({total - _TABLE_SHOWN_ROWS} rows skipped) ... |")
            continue
        md_lines.append("| " + " | ".join(cell.strip() for cell in row) + " |")
        if idx == 0:
            md_lines.append("| " + " | ".join("---" for _ in row) + " |")

    result = "\n".join(md_lines)
    if truncated:
        result += f"\n\n(Showing {_TABLE_SHOWN_ROWS} of {total} rows)"
    return result


def _collect_limited_rows(row_iter) -> tuple[list[list[str]], int]:
    """Стримит строки, не материализуя весь файл (защита от OOM на больших таблицах).

    Держит в памяти максимум head (_TABLE_HEAD_ROWS+1) + tail (_TABLE_TAIL_ROWS)
    строк. Если общее число строк <= порога усечения — возвращает все; иначе
    head + tail, чего достаточно для _rows_to_markdown. Второй элемент кортежа —
    точное общее число строк.

    Каждая ячейка приводится к str (Excel отдаёт значения разных типов).
    """
    from collections import deque

    head: list[list[str]] = []
    tail: deque[list[str]] = deque(maxlen=_TABLE_TAIL_ROWS)
    total = 0
    for raw in row_iter:
        total += 1
        row = [str(cell) if cell is not None else "" for cell in raw]
        if len(head) < _TABLE_READ_CAP:
            head.append(row)
        else:
            tail.append(row)

    if total <= _TABLE_TRUNCATE_THRESHOLD:
        return head, total
    # head хранит _TABLE_READ_CAP строк; _rows_to_markdown сам нарежет head+tail,
    # поэтому отдаём первые (_TABLE_HEAD_ROWS+1) + последние _TABLE_TAIL_ROWS.
    return head[: _TABLE_HEAD_ROWS + 1] + list(tail), total


def _read_csv(path: Path, encoding: str = "utf-8") -> str:
    """Reads a CSV/TSV file and returns content as a markdown table.

    For large files (>200 rows), shows first 100 and last 50 rows.
    """

    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","

    try:
        with open(path, encoding=encoding, errors="replace", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            rows, total = _collect_limited_rows(reader)
    except Exception as e:
        return f"[Error reading CSV: {e}]"

    if not rows:
        return "(empty file)"

    return _rows_to_markdown(rows, total=total)


def _read_excel(path: Path) -> str:
    """Reads an Excel file and returns all sheets as markdown tables."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "[Error: openpyxl not installed. Run: pip install openpyxl]"

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        return f"[Error opening Excel file: {e}]"

    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows, total = _collect_limited_rows(ws.iter_rows(values_only=True))

        if not rows:
            parts.append(f"## Sheet: {sheet_name}\n\n(empty sheet)")
            continue

        sheet_md = f"## Sheet: {sheet_name}\n\n" + _rows_to_markdown(rows, total=total)
        parts.append(sheet_md)

    wb.close()
    return "\n\n".join(parts)


def _read_docx(path: Path) -> str:
    """Reads DOCX through the native docx_engine in a compact block view.

    One output line equals one addressable Word block (bN).  This keeps normal
    `read` cheap even for long documents; exact run/format metadata is available
    on demand via the `docx` tool's inspect action.
    """
    from tools.file_ops.docx_tool import read_docx_compact

    return read_docx_compact(path)


def _read_pptx(path: Path) -> str:
    """Reads PPTX through the native pptx_engine in a compact slide view."""
    from tools.file_ops.pptx_tool import read_pptx_compact

    return read_pptx_compact(path)


def _read_pdf(path: Path) -> str:
    """Полное чтение PDF: текст (через PyMuPDF) + таблицы (через pdfplumber).

    Без лимитов на размер/число страниц. Возвращает markdown-подобный текст:
      - каждая страница с заголовком `## Page N`
      - текст в естественном порядке чтения (sort=True)
      - таблицы, найденные на странице, рендерятся ниже текста как markdown
    """
    try:
        import pymupdf  # PyMuPDF
    except ImportError:
        try:
            import fitz as pymupdf  # старое имя пакета
        except ImportError:
            return "[Error: pymupdf not installed. Run: uv add pymupdf]"

    try:
        pdfplumber = __import__("pdfplumber")
    except ImportError:
        pdfplumber = None

    try:
        doc = pymupdf.open(str(path))
    except Exception as e:
        return f"[Error opening PDF: {e}]"

    plumber_doc = None
    if pdfplumber is not None:
        try:
            plumber_doc = pdfplumber.open(str(path))
        except Exception:
            plumber_doc = None

    parts: list[str] = []
    meta = doc.metadata or {}
    title = (meta.get("title") or "").strip()
    author = (meta.get("author") or "").strip()
    header_bits = [f"pages: {doc.page_count}"]
    if title:
        header_bits.append(f"title: {title}")
    if author:
        header_bits.append(f"author: {author}")
    parts.append("[PDF · " + " · ".join(header_bits) + "]")

    for page_idx in range(doc.page_count):
        page = doc.load_page(page_idx)
        page_parts: list[str] = [f"## Page {page_idx + 1}"]

        try:
            text = page.get_text("text", sort=True) or ""
        except Exception as e:
            text = f"[text extraction failed: {e}]"
        text = text.strip()
        if text:
            page_parts.append(text)

        if plumber_doc is not None and page_idx < len(plumber_doc.pages):
            try:
                pl_page = plumber_doc.pages[page_idx]
                tables = pl_page.extract_tables() or []
            except Exception:
                tables = []
            for t_idx, table in enumerate(tables, start=1):
                if not table:
                    continue
                md_rows: list[str] = [f"\n### Table {page_idx + 1}.{t_idx}"]
                norm = [
                    [
                        (cell if cell is not None else "").replace("\n", " ").strip()
                        for cell in row
                    ]
                    for row in table
                ]
                if not norm or not norm[0]:
                    continue
                width = max(len(r) for r in norm)
                norm = [r + [""] * (width - len(r)) for r in norm]
                md_rows.append("| " + " | ".join(norm[0]) + " |")
                md_rows.append("| " + " | ".join("---" for _ in range(width)) + " |")
                for row in norm[1:]:
                    md_rows.append("| " + " | ".join(row) + " |")
                page_parts.append("\n".join(md_rows))

        parts.append("\n\n".join(page_parts))

    doc.close()
    if plumber_doc is not None:
        with contextlib.suppress(Exception):
            plumber_doc.close()

    return "\n\n".join(parts)


def _safe_read(path: Path, encoding: str = "utf-8") -> str:
    """Read a file fully."""
    return path.read_text(encoding=encoding, errors="replace")
